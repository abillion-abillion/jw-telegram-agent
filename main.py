"""
JW Financial Telegram Agent v4
- AGENTS.md 전역 컨텍스트 통합
- 대장/부장 사수-부사수 구조 유지
- 섹터 밸류에이션 분석 파이프라인 추가 (/분석, /전체분석)
- /stop, 인터럽트, 컨펌 흐름 유지
"""

import os
import io
import time
import base64
import httpx
import logging
import asyncio
from fastapi import FastAPI, Request
from contextlib import asynccontextmanager
from card_news_handler import handle_card_news_request, is_card_news_request

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ──────────────────────────────────────────
# 전역 상태
# ──────────────────────────────────────────
conversation_history: list[dict] = []
MAX_HISTORY = 20
pending_task: dict | None = None
is_working: bool = False
stop_requested: bool = False
interrupt_message: str | None = None

# ──────────────────────────────────────────
# 환경변수
# ──────────────────────────────────────────
TELEGRAM_TOKEN    = os.environ["TELEGRAM_BOT_TOKEN"]
TELEGRAM_API      = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"
ALLOWED_CHAT_ID   = int(os.environ["ALLOWED_CHAT_ID"])
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
DART_API_KEY      = os.environ.get("DART_API_KEY", "")
CLAUDE_MODEL      = "claude-sonnet-4-5"
GITHUB_TOKEN      = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO       = os.environ.get("GITHUB_REPO", "abillion-abillion/jw-agent-workspace")

# ──────────────────────────────────────────
# JW대장 시스템 프롬프트 (AGENTS.md 완전 통합)
# ──────────────────────────────────────────

DAEJANG_PROMPT = """
당신은 JW대장입니다. JW Financial Consulting의 AI 사수 에이전트입니다.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
## 운영자 정보 (SSoT)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- 이름: 남진우 | 자산관리사/팀장 5년차 | 하나증권 연계 JW Financial Consulting
- 사이트: jwfinancial.co.kr (GitHub Pages, PWA)
- 리서치 브랜드: 핀사이트랩스 (FINSIGHT LABS) — 매크로 분석, 기관투자자 시각
- 블로그: 네이버 블로그 moneymustard | 유튜브: @honeymoneymustard
- 주요 채널: KakaoTalk (Solapi 비즈채널), Telegram

## 사업 구조
- 주 수입원: 컨설팅 비용(Cal.com 예약), 금융계약(변액연금/IRP/ISA/보장성보험)
- 부 수입원: 교육콘텐츠/지식창업, 핀사이트랩스 리서치
- 핵심 타겟: 30대 맞벌이 직장인 | 총 클라이언트 100~150명
- 확장 타겟: 네이버 직원, 중국어권 서울 거주자 (위챗/샤오홍슈)
- 팀: 진우님(대표) + 주니어 1명(한국어/중국어 이중언어)
- 금지 단어: 가입, 계약, 납입, 판매 (고객 발송물 전체 적용)

## 운영중인 자동화 파이프라인
- 모닝 브리핑: RSS → Claude API → Telegram ✅
- 상담 분석: Notion DB 폴링 → Claude/Gemini → Telegram ✅
- 상담 인테이크: Google Forms → GAS → Telegram ✅
- Clova Note 요약: 녹취 → Claude → 상담요약 + 카카오초안 ✅
- 섹터 밸류에이션: pykrx/DART → Claude → Telegram ✅

## 현재 진행중 프로젝트
- 라에미안 라그란데 등기이전 (관리처분 변경 총회 5~6월, 등기 9~10월 목표)
- 보험료 비교 시뮬레이터 (기획 단계)
- IFP vs AFPK 자격증 시험 2026-05-16
- 네이버 직원 대상 마케팅 카드이미지 시리즈

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
## 역할
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- 진우님 업무 지시 → JW부장(부사수)에게 작업 지시
- JW부장 결과물 검토 및 품질 관리
- 최종본 진우님 보고
- 자유 질문에는 직접 답변

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
## 스킬 라우팅 — 작업 유형 자동 판단
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

진우님 요청을 받으면 아래 기준으로 스킬을 먼저 판단하고,
JW부장에게 지시할 때 해당 스킬의 프레임워크를 반드시 명시한다.
복합 요청은 여러 스킬을 조합해서 지시한다.

### 1. SALES 스킬 (skills/sales.md)
트리거: 상담, 미팅, 1차/2차/3차, 클로징, 제안서, 거절, 이탈방지, 소개요청, 레퍼럴, 신규고객
프레임워크:
- 대화/스크립트 → SPIN Selling 4단계 (S→P→I→N 순서 엄수)
- 제안서/가치 설계 → Hormozi 가치 방정식 (결과↑ 가능성↑ 시간↓ 노력↓) + 5요소 구조
- 거절/저항 대응 → Tracy 심리 패턴 ("생각해볼게요" / "상의해야" / "여유없어" 각 대응)
검토 기준:
- SPIN 순서 틀림 → 순서 교정 지시
- Hormozi 5요소 누락 → 누락 요소 추가 지시
- 고객이 스스로 결론 내리는 구조인가

### 2. MARKETING 스킬 (skills/marketing.md)
트리거: 카드뉴스, 블로그, SNS, 카카오채널, 숏폼, 유튜브, 핀사이트랩스, 광고, 브랜딩, 콘텐츠 전략
프레임워크:
- 브랜드 판단 먼저 → JW Financial(실용/친근) vs 핀사이트랩스(분석/기관 시각) vs 확장채널(중국어권)
- 채널별 형식 맞춤 → 네이버(SEO+정보) / 카카오(3~5줄+CTA) / 유튜브(후킹 인트로) / 숏폼(첫 3초) / 샤오홍슈(생활감)
- 독자 인식 레벨 → Schwartz 5단계 (무인식→문제→해결책→상품→구매 준비)
- 콘텐츠 믹스 → Gary V 4:1 (가치 4 : 전환 1)
검토 기준:
- Purple Cow 체크리스트 통과 (경쟁 콘텐츠와 30% 이상 다른가)
- CTA 1개 이하인가
- 모바일 첫 2줄에 핵심 전달되는가

### 3. FINANCE 스킬 (skills/finance-consulting.md)
트리거: 포트폴리오, 시뮬레이션, 수익률, 자산배분, 매크로, 재무진단, 노후설계,
        IRP, ISA, 핀사이트랩스 리포트, PER, PBR, EV/EBITDA, 섹터분석, 밸류에이션
프레임워크:
- 고객 재무진단 → Damodaran 5단계 (현금흐름→자산부채→목표역산→갭분석→포트폴리오)
- 매크로 분석 → Howard Marks 사이클 포지션 + 핀사이트랩스 5단계 리포트 구조
- 상품/개념 설명 → Buffett/Munger 비유 라이브러리 + 역산 사고
- 섹터 밸류에이션 → PBR<1 + PER<15 + EV/EBITDA<10 저평가 신호 / 업종 평균 대비 프리미엄/디스카운트
검토 기준:
- 수익률은 반드시 3시나리오 (보수 3~4% / 중립 5~6% / 낙관 7~8%)
- 보장성·확정 수익 표현 금지 → ⚠️ 확인필요 표기 지시
- 비유가 먼저 나왔는가
- 투자 권유 언어 금지 — 모든 금융 분석은 참고용임을 명시

### 4. CONTENT CREATION 스킬 (skills/content-creation.md)
트리거: 써줘, 초안, 작성해줘, 카드뉴스 본문, 블로그 본문, 스크립트, 안내문, 공지, 이메일 초안
※ MARKETING 스킬과 함께 사용 (marketing = 전략/방향, content-creation = 실제로 쓰는 법)
프레임워크:
- 구조 설계 → StoryBrand 7요소 (고객이 주인공, JW는 가이드)
- 문장 원칙 → Handley TREAT (진실/관련/공감/적절/타겟)
- 아이디어 확장 → Perell 원자→분자 전략 (1 아이디어 → 5채널 재활용)
검토 기준:
- 고객(독자)이 주인공인가 (브랜드 자랑이 주인공 아닌가)
- 첫 문장이 스크롤을 멈추게 하는가
- 진우님 관점/경험이 담겼는가 (단순 정보 나열 금지)

### 5. KAKAO MESSAGE 스킬 (skills/kakao-message.md)
트리거: 카카오, 카톡, 문자, 2차완료, 계약완료, 온보딩, 이탈방지, 팔로업, 소개요청, 시퀀스
프레임워크:
- 고객 여정 단계 판단 → Coleman 8단계
- 시퀀스 선택 →
    B. 이탈방지 (2차완료 → D+1/2/3)
    C. 온보딩 (계약완료 → D+0/7/30/90)
    D. 소개 요청 (D+90 이후, Give 먼저 — 재무진단 보고서/핀사이트랩스 리포트/체크리스트 키트)
- 단발 메시지 → A-1~A-8 템플릿 중 선택
검토 기준:
- 금지 단어 없는가 (가입/계약/납입/판매)
- 이름 호명, 5줄 이내, CTA 1개, 이모지 2개 이하
- ✅ 검토 후 발송 표기 / ⚠️ 확인필요 항목 명시

### 6. OPERATIONS 스킬 (skills/operations.md)
트리거: 자동화, 파이프라인, 봇, 노션, CRM, 프로세스, SOP, 루틴, 병목, 주간리뷰, 대시보드, 코드, 개발
프레임워크:
- 업무 수집/정리 → GTD 5단계 + 주간 리뷰 체크리스트
- 낭비 제거 → 도요타 7가지 낭비 진단
- 병목 개선 → TOC 5단계 분석 (현재→병목→원인→개선→효과)
- 신규 자동화 → 기존 인프라 기반 설계 (FastAPI/Python/Railway/Telegram)
검토 기준:
- 자동화 가능 vs 진우님 직접 처리 명확히 구분됐는가
- TRIGGER→ACTION→CONDITION→OUTPUT→HUMAN 구조 완성됐는가

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
## 스킬 복합 적용 가이드
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

카드뉴스 초안 요청
  → MARKETING (채널/독자/각도) + CONTENT CREATION (실제 작성)

2차완료 [고객명]
  → KAKAO MESSAGE (이탈방지 D+1/2/3) + SALES (이탈방지 논리 보강)

포트폴리오 북 [고객명]
  → FINANCE (재무진단 6섹션) + CONTENT CREATION (작성 원칙)

핀사이트랩스 리포트
  → FINANCE (매크로/밸류에이션 분석) + MARKETING (핀사이트랩스 톤) + CONTENT CREATION (작성)

자동화 설계 요청
  → OPERATIONS (파이프라인 설계) + 관련 스킬 (어떤 업무인지에 따라 조합)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
## 업무 처리 방식
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. 진우님 요청 분석 → 스킬 라우팅 판단 (복합 요청 시 조합)
2. JW부장에게 구체적 지시 ('JW부장에게:'로 시작, 적용 스킬 + 프레임워크 명시)
3. JW부장 초안 검토:
   - 해당 스킬 검토 기준 적용
   - 고객 발송용: ✅ 검토 후 발송 포함 여부
   - 금융 수치: ⚠️ 확인필요 표기 여부
   - 금지 단어 (가입/계약/납입/판매) 없는지
4. 수정 필요시 구체적 재지시 (어느 부분, 왜, 어떻게)
5. OK면 "✅ 대장 검토 완료" 후 진우님 보고

## 콘텐츠 원칙
- 어려운 금융/경제를 쉽게: 비유 먼저, 인과관계 중심
- 짧고 명확하게: 카드뉴스/숏폼 친화적, 모바일 가독성
- 행동 유도: 고객용 콘텐츠에 상담 CTA 포함
- 수치/날짜 공백 → [○○] + ⚠️ 확인필요 표기

⚠️ 투자 권유 언어 사용 금지 — 모든 금융 분석은 참고용임을 명시
"""

# ──────────────────────────────────────────
# JW부장 시스템 프롬프트
# ──────────────────────────────────────────
BUJANG_PROMPT = """
당신은 JW부장입니다. JW Financial Consulting의 AI 부사수 에이전트입니다.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
역할
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

JW대장 지시를 받아 실제 작업물(초안) 생성
검토 피드백 반영하여 빠르게 수정
수정 반영 시 변경 부분 앞에 ▶ 표시

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
운영자 컨텍스트
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

이름: 남진우 | 자산관리사/팀장 5년차 | 하나증권 연계 JW Financial
핵심 고객: 30대 맞벌이 직장인 | 클라이언트 100~150명
주력 상품: 변액연금(10년납), 보장성보험, IRP, ISA, 연금저축펀드, 달러연금보험
리서치 브랜드: 핀사이트랩스 (매크로 분석, 기관투자자 시각)
이탈 최다 구간: 2차→3차 사이

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
절대 원칙 (모든 작업 공통)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

대장 지시사항 빠짐없이 반영
금지 단어: 가입, 계약, 납입, 판매 (고객 발송물 전체)
고객 발송용: "✅ 검토 후 발송해주세요" 필수
수치/날짜 불확실: [○○] + "⚠️ 확인필요" 명시
판매 언어 금지 → 교육/공감 톤 유지
수정 반영 시 변경 부분 앞에 ▶ 표시
투자 권유 언어 금지 → 참고용임을 명시

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
스킬별 작업 원칙
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SALES (skills/sales.md)

SPIN 질문 순서 엄수: S→P→I→N (순서 바꾸지 않음)
Hormozi 제안서: 꿈의 결과 + 갭 + 지금 이유 + JW 이유 + 리스크 제거
Tracy 거절 대응: "생각해볼게요" → "어떤 부분이 걸리세요?" 직접 질문
고객이 스스로 결론 내리는 구조로 작성

MARKETING (skills/marketing.md)

브랜드 먼저 확인: JW Financial / 핀사이트랩스 / 중국어권
채널별 형식 준수:
네이버 블로그: 800~1500자, 목차, SEO 키워드
카카오채널: 3~5줄, 이모지 절제, CTA 1개
숏폼: 첫 3초 후킹 멘트 먼저
Schwartz 인식 레벨 맞는 메시지 작성
Purple Cow 체크: 경쟁 콘텐츠와 30% 이상 다른가

FINANCE (skills/finance-consulting.md)

수익률: 반드시 3시나리오 (보수 3~4% / 중립 5~6% / 낙관 7~8%)
보장성 표현 절대 금지
비유 먼저, 숫자 나중
포트폴리오 북: 재무현황/SWOT/목표/포트폴리오/상품설명/로드맵 6섹션
핀사이트랩스 리포트: 한줄요약→현황→해석→사이클→시사점→액션

CONTENT CREATION (skills/content-creation.md)

StoryBrand: 고객이 주인공, JW Financial은 가이드
첫 문장 공식: 질문형 / 상황형 / 숫자형 / 역설형 중 선택
단순 정보 나열 금지 → 진우님 관점/경험 반드시 포함
콘텐츠 품질 체크리스트 8항목 자가검토 후 제출

KAKAO MESSAGE (skills/kakao-message.md)

이름 호명 필수, 5줄 이내, CTA 1개, 이모지 2개 이하
시퀀스 유형 확인:
2차완료 → B. 이탈방지 (D+1/2/3)
계약완료 → C. 온보딩 (D+0/7/30/90)
소개요청 → D. Give 먼저 (재무진단/리포트/체크리스트)
✅ 검토 후 발송 + ⚠️ 확인필요 항목 반드시 표기

OPERATIONS (skills/operations.md)

자동화 설계: TRIGGER→ACTION→CONDITION→OUTPUT→HUMAN 구조
낭비 제거: 7가지 낭비 중 해당 항목 먼저 파악
자동화 가능 vs 진우님 직접 처리 명확히 구분

CS/CRM (skills/cs-crm.md)

고객 3티어 분류: Promoter / Passive / Detractor
이탈 위험 신호: 2회 이상 무응답 / 3개월 이상 접촉 없음
불만 고객: 즉시 기록 → 사실 파악 → 해결 → 사후 확인 순서

HR (skills/hr.md)

채용 기준: 태도 > 지식 (금융 지식은 교육 가능)
온보딩: 30/60/90일 단계별 설계
피드백: 잘한 것 1개 + 개선 1개 + 다음 목표 1개 (15분 구조)

LEGAL (skills/legal.md)

모든 법률 정보는 "참고용" 명시, 전문가 상담 권고
콘텐츠 금지: "보장", "확정", "반드시" 표현
고객 발송 전: 금융 광고 컴플라이언스 체크리스트 적용

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
자가검토 체크리스트 (제출 전)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
□ 대장 지시사항 모두 반영됐는가
□ 금지 단어 (가입/계약/납입/판매) 없는가
□ 고객 발송용 — ✅ 검토 후 발송 표기됐는가
□ 불확실 수치 — ⚠️ 확인필요 표기됐는가
□ 해당 스킬 핵심 원칙 지켰는가
□ 첫 문장이 독자를 멈추게 하는가 (콘텐츠)
□ 판매 언어 없는가
□ 투자 권유 언어 없는가 (금융 콘텐츠)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
자주 쓰는 비유 라이브러리
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
복리:       "눈덩이 굴리기 — 경사가 길수록 커진다"
보험:       "소방서 — 쓸 일 없길 바라지만 없으면 끝"
IRP:        "세금 환급 전용 통장"
분산투자:   "한 바구니에 달걀 다 담지 않기"
변액연금:   "10년 씨앗, 그다음은 나무가 알아서"
인플레이션: "10년 전 짜장면 값을 생각해보세요"
사이클:     "경기도 봄여름가을겨울이 있다"
자산배분:   "계절별로 옷 갈아입듯 포트폴리오도"
"""

# ──────────────────────────────────────────
# 작업 요청 키워드
# ──────────────────────────────────────────
TASK_KEYWORDS = [
    "만들어줘", "작성해줘", "짜줘", "초안", "써줘",
    "포트폴리오 북", "이탈방지", "온보딩", "소개 요청",
    "카드뉴스", "블로그", "리포트", "메시지", "시퀀스",
    "2차완료", "계약완료", "분석해줘", "정리해줘"
]

# 섹터분석은 별도 파이프라인
SECTOR_KEYWORDS = ["섹터분석", "밸류에이션", "PER 분석", "PBR 분석"]

# ──────────────────────────────────────────
# FastAPI
# ──────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    await set_webhook()
    yield

app = FastAPI(lifespan=lifespan)


# ──────────────────────────────────────────
# Telegram 유틸
# ──────────────────────────────────────────
async def set_webhook():
    webhook_url = os.environ.get("WEBHOOK_URL", "")
    if not webhook_url:
        return
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{TELEGRAM_API}/setWebhook",
            json={"url": f"{webhook_url}/webhook", "drop_pending_updates": True}
        )
        logger.info(f"Webhook 등록: {resp.json()}")


async def send_message(chat_id: int, text: str):
    chunks = [text[i:i+4000] for i in range(0, len(text), 4000)]
    async with httpx.AsyncClient() as client:
        for chunk in chunks:
            await client.post(
                f"{TELEGRAM_API}/sendMessage",
                json={"chat_id": chat_id, "text": chunk, "parse_mode": "Markdown"}
            )


async def send_typing(chat_id: int):
    async with httpx.AsyncClient() as client:
        await client.post(
            f"{TELEGRAM_API}/sendChatAction",
            json={"chat_id": chat_id, "action": "typing"}
        )


async def download_file(file_id: str) -> bytes:
    async with httpx.AsyncClient() as client:
        resp = await client.get(f"{TELEGRAM_API}/getFile?file_id={file_id}")
        file_path = resp.json()["result"]["file_path"]
        file_resp = await client.get(
            f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}"
        )
        return file_resp.content


# ──────────────────────────────────────────
# Claude API
# ──────────────────────────────────────────
async def call_claude(system: str, messages: list, max_tokens: int = 2000) -> str:
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json"
    }
    async with httpx.AsyncClient(timeout=120.0) as client:
        resp = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers=headers,
            json={
                "model": CLAUDE_MODEL,
                "max_tokens": max_tokens,
                "system": system,
                "messages": messages
            }
        )
        resp.raise_for_status()
        return resp.json()["content"][0]["text"]


# ──────────────────────────────────────────
# 중단/인터럽트 체크
# ──────────────────────────────────────────
async def check_stop(chat_id: int) -> bool:
    global stop_requested, is_working
    if stop_requested:
        stop_requested = False
        is_working = False
        await send_message(chat_id, "⛔ 작업이 중단됐습니다.")
        return True
    return False


async def check_interrupt(chat_id: int) -> bool:
    global interrupt_message, is_working
    if interrupt_message:
        msg = interrupt_message
        interrupt_message = None
        await send_message(chat_id, f"⚡ *긴급 처리 중...*\n`{msg}`")
        await send_typing(chat_id)
        reply = await call_claude(
            system=DAEJANG_PROMPT,
            messages=[{"role": "user", "content": msg}],
            max_tokens=800
        )
        await send_message(chat_id, f"💬 *대장 답변*\n\n{reply}")
        await send_message(chat_id,
            "━━━━━━━━━━━━━━━━━━━━\n"
            "진행 중인 작업을 계속할까요?\n"
            "• *'계속'* — 작업 재개\n"
            "• */stop* — 작업 중단"
        )
        for _ in range(60):
            await asyncio.sleep(1)
            if stop_requested:
                stop_requested = False
                is_working = False
                await send_message(chat_id, "⛔ 작업이 중단됐습니다.")
                return True
            if interrupt_message == "계속":
                interrupt_message = None
                await send_message(chat_id, "▶️ 작업을 재개합니다...")
                return False
        await send_message(chat_id, "⏱ 응답 없음 — 작업을 자동 재개합니다...")
        return False
    return False


# ──────────────────────────────────────────
# 섹터 밸류에이션 파이프라인
# ──────────────────────────────────────────
# 설정값 (환경변수 or 기본값)
ANALYSIS_SECTORS = os.environ.get("TARGET_SECTORS", "반도체").split(",")
ANALYSIS_MARKET_CAP_MIN = 100_000_000_000
ANALYSIS_TOP_N = 15
ANALYSIS_YEAR = "2024"
ANALYSIS_REPRT = "11011"


def _get_last_business_day() -> str:
    from datetime import datetime, timedelta
    d = datetime.today() - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d.strftime("%Y%m%d")


def _fetch_sector_data_sync(sector_name: str) -> "pd.DataFrame | None":
    """pykrx → KRX fallback으로 섹터 데이터 수집 (동기)"""
    try:
        import pandas as pd
        from pykrx import stock as krx

        date = _get_last_business_day()
        sector_df      = krx.get_market_sector_classifications(date, market="KOSPI")
        sector_tickers = sector_df[sector_df["섹터"] == sector_name].index.tolist()
        if not sector_tickers:
            return None

        cap_df         = krx.get_market_cap_by_ticker(date, market="KOSPI")
        fundamental_df = krx.get_market_fundamental_by_ticker(date, market="KOSPI")

        rows = []
        for t in sector_tickers:
            if t not in cap_df.index or t not in fundamental_df.index:
                continue
            rows.append({
                "ticker":     t,
                "name":       krx.get_market_ticker_name(t),
                "market_cap": cap_df.loc[t, "시가총액"],
                "per":        fundamental_df.loc[t, "PER"],
                "pbr":        fundamental_df.loc[t, "PBR"],
                "eps":        fundamental_df.loc[t, "EPS"],
                "bps":        fundamental_df.loc[t, "BPS"],
            })

        df = pd.DataFrame(rows)
        df = df[df["market_cap"] >= ANALYSIS_MARKET_CAP_MIN]
        df = df[df["per"] > 0]
        df = df[df["pbr"] > 0]
        return df.sort_values("market_cap", ascending=False).head(ANALYSIS_TOP_N).reset_index(drop=True)

    except Exception as e:
        logger.error(f"섹터 데이터 수집 실패: {e}")
        return None


def _fetch_dart_sync(tickers: list) -> dict:
    """DART OpenAPI → EV/EBITDA용 재무 데이터 수집 (동기)"""
    if not DART_API_KEY:
        return {}

    import zipfile, io as _io, xml.etree.ElementTree as ET, requests as req

    # corp_code 맵
    try:
        r = req.get("https://opendart.fss.or.kr/api/corpCode.xml",
                    params={"crtfc_key": DART_API_KEY}, timeout=15)
        z = zipfile.ZipFile(_io.BytesIO(r.content))
        root = ET.fromstring(z.read("CORPCODE.xml"))
        corp_map = {
            item.findtext("stock_code", "").strip(): item.findtext("corp_code", "").strip()
            for item in root.findall("list")
            if item.findtext("stock_code", "").strip()
        }
    except Exception as e:
        logger.error(f"DART corp_code 로드 실패: {e}")
        return {}

    result = {}
    for ticker in tickers:
        corp_code = corp_map.get(ticker)
        if not corp_code:
            continue
        try:
            for fs_div in ("CFS", "OFS"):
                r = req.get(
                    "https://opendart.fss.or.kr/api/fnlttSinglAcnt.json",
                    params={"crtfc_key": DART_API_KEY, "corp_code": corp_code,
                            "bsns_year": ANALYSIS_YEAR, "reprt_code": ANALYSIS_REPRT,
                            "fs_div": fs_div},
                    timeout=10
                )
                data = r.json()
                if data.get("status") == "000":
                    break

            accounts = {item["account_nm"]: item for item in data.get("list", [])}

            def _parse(name):
                item = accounts.get(name)
                if item:
                    try: return int(item.get("thstrm_amount", "0").replace(",", ""))
                    except: return None
                return None

            result[ticker] = {
                "operating_profit": _parse("영업이익"),
                "depreciation":     _parse("감가상각비"),
                "total_debt":       _parse("부채총계"),
            }
        except Exception as e:
            logger.error(f"DART {ticker} 실패: {e}")
        time.sleep(0.3)

    return result


def _calc_ev_ebitda(market_cap, total_debt, operating_profit, depreciation):
    if None in (market_cap, total_debt, operating_profit, depreciation):
        return None
    ebitda = operating_profit + depreciation
    if ebitda <= 0:
        return None
    return round((market_cap + total_debt) / ebitda, 2)


def _valuation_signal(row) -> str:
    import pandas as pd
    score, flags = 0, []
    if pd.notna(row.get("pbr"))      and row["pbr"] < 1.0:   score += 1; flags.append("PBR<1")
    if pd.notna(row.get("per"))      and row["per"] < 15:     score += 1; flags.append("PER<15")
    if pd.notna(row.get("ev_ebitda")) and row["ev_ebitda"] < 10: score += 1; flags.append("EV/EBITDA<10")
    if score == 3: return f"🟢 저평가 ({', '.join(flags)})"
    if score == 2: return f"🟡 관심 ({', '.join(flags)})"
    if score == 1: return f"⚪ 보통 ({', '.join(flags)})"
    return "🔴 고평가"


async def run_sector_analysis(chat_id: int, sector_name: str):
    """섹터 밸류에이션 전체 파이프라인 → Telegram 발송"""
    import pandas as pd
    from datetime import datetime

    await send_message(chat_id, f"⏳ *{sector_name}* 섹터 분석 시작... (1~2분 소요)")
    loop = asyncio.get_event_loop()

    # 1. 데이터 수집 (blocking → executor)
    df = await loop.run_in_executor(None, _fetch_sector_data_sync, sector_name)
    if df is None or df.empty:
        await send_message(chat_id, f"❌ *{sector_name}* 섹터 종목을 찾을 수 없습니다.\n섹터명 확인: /섹터목록")
        return

    # 2. DART 재무 수집
    dart_data = await loop.run_in_executor(None, _fetch_dart_sync, df["ticker"].tolist())

    # 3. EV/EBITDA 계산
    rows = []
    for _, row in df.iterrows():
        d = dart_data.get(row["ticker"], {})
        ev_ebitda = _calc_ev_ebitda(
            row["market_cap"], d.get("total_debt") or 0,
            d.get("operating_profit") or 0, d.get("depreciation") or 0
        )
        rows.append({**row.to_dict(), "ev_ebitda": ev_ebitda})
    enriched = pd.DataFrame(rows)
    enriched["signal"] = enriched.apply(_valuation_signal, axis=1)

    # 4. Claude AI 코멘트
    summary_rows = []
    for _, r in enriched.iterrows():
        ev = f"{r['ev_ebitda']:.1f}" if pd.notna(r.get("ev_ebitda")) else "N/A"
        summary_rows.append(f"- {r['name']}({r['ticker']}): PER {r['per']:.1f}, PBR {r['pbr']:.2f}, EV/EBITDA {ev}")

    comment = await call_claude(
        system=DAEJANG_PROMPT,
        messages=[{"role": "user", "content":
            f"{sector_name} 섹터 밸류에이션 데이터:\n" + "\n".join(summary_rows) +
            "\n\n200자 이내로: ①섹터 전반 밸류에이션 수준 ②주목할 종목 1~2개 ③투자자 유의사항. 핀사이트랩스 톤앤매너로."}],
        max_tokens=400
    )

    # 5. 메시지 포맷
    date_str = datetime.today().strftime("%Y-%m-%d")
    lines = [
        f"📊 *{sector_name} 섹터 밸류에이션*",
        f"_기준일: {date_str} | 출처: KRX/DART_",
        "",
        "```",
        f"{'종목':<8} {'PER':>5} {'PBR':>5} {'EV/EB':>6}  신호",
        "─" * 38,
    ]
    for _, r in enriched.iterrows():
        ev  = f"{r['ev_ebitda']:.1f}" if pd.notna(r.get("ev_ebitda")) else "  N/A"
        sig = "🟢" if "저평가" in r["signal"] else "🟡" if "관심" in r["signal"] else "⚪" if "보통" in r["signal"] else "🔴"
        lines.append(f"{r['name'][:6]:<8} {r['per']:>5.1f} {r['pbr']:>5.2f} {ev:>6}  {sig}")
    lines += [
        "```", "",
        "🤖 *핀사이트랩스 코멘트*",
        comment, "",
        "⚠️ _투자 참고용 — 투자 권유 아님_"
    ]

    await send_message(chat_id, "\n".join(lines))


# ──────────────────────────────────────────
# 사수-부사수 파이프라인 (기존 유지)
# ──────────────────────────────────────────
def is_task_request(text: str) -> bool:
    return any(k in text for k in TASK_KEYWORDS)


async def run_team_pipeline(chat_id: int, user_request: str, file_blocks: list = None):
    global is_working, stop_requested, pending_task
    is_working = True
    stop_requested = False

    await send_message(chat_id,
        "⚙️ *대장-부장 팀 작업 시작*\n"
        "작업 중단: `/stop`\n"
        "긴급 질문: 바로 메시지 보내세요"
    )

    try:
        # STEP 1: 대장 → 부장 지시
        if await check_stop(chat_id): return
        await send_typing(chat_id)
        daejang_instruction = await call_claude(
            system=DAEJANG_PROMPT,
            messages=[{"role": "user", "content":
                f"진우님 요청:\n{user_request}\n\n"
                "JW부장에게 구체적으로 지시해주세요. 'JW부장에게:'로 시작하세요."}],
            max_tokens=800
        )
        if await check_stop(chat_id): return
        await send_message(chat_id, f"📋 *대장 → 부장 지시*\n\n{daejang_instruction}")

        # STEP 2: 부장 초안
        if await check_stop(chat_id): return
        if await check_interrupt(chat_id): return
        await send_message(chat_id, "✏️ *JW부장 초안 작성 중...*")
        await send_typing(chat_id)

        bujang_content = f"대장 지시:\n{daejang_instruction}\n\n진우님 요청:\n{user_request}"
        bujang_messages = [{"role": "user", "content":
            (file_blocks + [{"type": "text", "text": bujang_content}]) if file_blocks else bujang_content}]

        draft = await call_claude(system=BUJANG_PROMPT, messages=bujang_messages, max_tokens=2500)
        if await check_stop(chat_id): return
        await send_message(chat_id, f"📝 *JW부장 초안*\n\n{draft}")

        # STEP 3: 대장 검토
        if await check_stop(chat_id): return
        if await check_interrupt(chat_id): return
        await send_message(chat_id, "🔍 *대장 검토 중...*")
        await send_typing(chat_id)

        review = await call_claude(
            system=DAEJANG_PROMPT,
            messages=[{"role": "user", "content":
                f"진우님 요청:\n{user_request}\n\nJW부장 초안:\n{draft}\n\n"
                "검토해주세요. 수정 필요시 구체적으로 지시, OK면 '✅ 대장 검토 완료'로 마무리."}],
            max_tokens=800
        )
        if await check_stop(chat_id): return
        await send_message(chat_id, f"🔎 *대장 검토*\n\n{review}")

        # STEP 3-2: 수정 라운드
        if "검토 완료" not in review and "✅" not in review:
            if await check_stop(chat_id): return
            if await check_interrupt(chat_id): return
            await send_message(chat_id, "🔄 *JW부장 수정 중...*")
            await send_typing(chat_id)
            revised = await call_claude(
                system=BUJANG_PROMPT,
                messages=[
                    {"role": "user", "content": f"초안:\n{draft}"},
                    {"role": "assistant", "content": draft},
                    {"role": "user", "content": f"대장 피드백:\n{review}\n\n반영해서 수정. 변경 부분 앞에 ▶ 표시."}
                ],
                max_tokens=2500
            )
            if await check_stop(chat_id): return
            await send_message(chat_id, f"📝 *JW부장 수정본*\n\n{revised}")

            final_review = await call_claude(
                system=DAEJANG_PROMPT,
                messages=[{"role": "user", "content": f"수정본:\n{revised}\n\n최종 검토해주세요."}],
                max_tokens=400
            )
            if await check_stop(chat_id): return
            await send_message(chat_id, f"✅ *대장 최종 검토*\n\n{final_review}")
            draft = revised

        # STEP 4: 컨펌 요청
        pending_task = {"original": user_request, "draft": draft}
        is_working = False
        await send_message(chat_id,
            "━━━━━━━━━━━━━━━━━━━━\n"
            "📌 *진우님 컨펌 요청*\n\n"
            "• *'확정'* — 그대로 사용\n"
            "• *'수정: [내용]'* — 수정 요청\n"
            "• *'다시'* — 처음부터 재작업"
        )

    except asyncio.CancelledError:
        is_working = False
        await send_message(chat_id, "⛔ 작업이 중단됐습니다.")
    except Exception as e:
        is_working = False
        logger.error(f"파이프라인 오류: {e}")
        await send_message(chat_id, f"⚠️ 작업 중 오류: {str(e)}")


# ──────────────────────────────────────────
# 컨펌 처리 (기존 유지)
# ──────────────────────────────────────────
async def handle_confirm(chat_id: int, text: str) -> bool:
    global pending_task
    if not pending_task:
        return False

    if text.strip() == "확정":
        await send_message(chat_id, "✅ 확정됐습니다! 다음 업무 말씀해주세요.")
        pending_task = None
        return True

    if text.strip() == "다시":
        original = pending_task["original"]
        pending_task = None
        await send_message(chat_id, "🔄 처음부터 다시 작업합니다...")
        asyncio.create_task(run_team_pipeline(chat_id, original))
        return True

    if text.startswith("수정:"):
        revision_note = text[3:].strip()
        original_draft = pending_task["draft"]
        await send_message(chat_id, f"🔄 *수정 반영 중...*\n요청: {revision_note}")
        await send_typing(chat_id)

        revised = await call_claude(
            system=BUJANG_PROMPT,
            messages=[
                {"role": "user", "content": f"기존 작업물:\n{original_draft}"},
                {"role": "assistant", "content": original_draft},
                {"role": "user", "content": f"진우님 수정 요청:\n{revision_note}\n\n수정해주세요. 변경 부분 앞에 ▶ 표시."}
            ],
            max_tokens=2500
        )
        review = await call_claude(
            system=DAEJANG_PROMPT,
            messages=[{"role": "user", "content":
                f"진우님 수정 요청: {revision_note}\n\n수정본:\n{revised}\n\n검토해주세요."}],
            max_tokens=400
        )
        pending_task["draft"] = revised
        await send_message(chat_id, f"📝 *수정본*\n\n{revised}")
        await send_message(chat_id, f"🔎 *대장 검토*\n\n{review}")
        await send_message(chat_id,
            "━━━━━━━━━━━━━━━━━━━━\n"
            "📌 *진우님 컨펌 요청*\n\n"
            "• *'확정'* / *'수정: [내용]'* / *'다시'*"
        )
        return True

    return False


# ──────────────────────────────────────────
# 단순 대화 (대장 직접)
# ──────────────────────────────────────────
async def ask_daejang(user_message: str, file_blocks: list = None) -> str:
    global conversation_history
    content = (file_blocks + [{"type": "text", "text": user_message or "분석해줘"}]) if file_blocks else user_message
    conversation_history.append({"role": "user", "content": content})
    if len(conversation_history) > MAX_HISTORY:
        conversation_history = conversation_history[-MAX_HISTORY:]
    reply = await call_claude(system=DAEJANG_PROMPT, messages=conversation_history, max_tokens=1500)
    conversation_history.append({"role": "assistant", "content": reply})
    return reply


# ──────────────────────────────────────────
# 파일 파싱 (기존 유지)
# ──────────────────────────────────────────
async def parse_file(file_bytes: bytes, mime_type: str, file_name: str) -> dict:
    mime = (mime_type or "").lower()
    name = (file_name or "").lower()

    if mime.startswith("image/") or name.endswith((".jpg", ".jpeg", ".png", ".webp")):
        img_mime = "image/jpeg" if name.endswith((".jpg", ".jpeg")) else "image/png"
        return {"type": "image", "source": {
            "type": "base64", "media_type": img_mime,
            "data": base64.standard_b64encode(file_bytes).decode("utf-8")}}

    if mime == "application/pdf" or name.endswith(".pdf"):
        return {"type": "document", "source": {
            "type": "base64", "media_type": "application/pdf",
            "data": base64.standard_b64encode(file_bytes).decode("utf-8")}}

    if name.endswith(".docx"):
        try:
            import docx
            doc = docx.Document(io.BytesIO(file_bytes))
            text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
            return {"type": "text", "text": f"[Word: {file_name}]\n\n{text}"}
        except Exception as e:
            return {"type": "text", "text": f"[Word 파싱 실패: {e}]"}

    if name.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            lines = []
            for sheet in wb.sheetnames[:3]:
                ws = wb[sheet]
                lines.append(f"[시트: {sheet}]")
                for row in ws.iter_rows(max_row=100, values_only=True):
                    row_text = "\t".join([str(c) if c is not None else "" for c in row])
                    if row_text.strip():
                        lines.append(row_text)
            return {"type": "text", "text": f"[Excel: {file_name}]\n\n" + "\n".join(lines)}
        except Exception as e:
            return {"type": "text", "text": f"[Excel 파싱 실패: {e}]"}

    try:    text = file_bytes.decode("utf-8")
    except: text = file_bytes.decode("cp949", errors="replace")
    return {"type": "text", "text": f"[파일: {file_name}]\n\n{text}"}


async def handle_file_message(chat_id: int, message: dict):
    await send_typing(chat_id)
    caption = message.get("caption", "이 파일을 분석해줘")
    file_id, mime_type, file_name = None, "", ""

    if "document" in message:
        doc = message["document"]
        file_id, mime_type = doc["file_id"], doc.get("mime_type", "")
        file_name = doc.get("file_name", "파일")
    elif "photo" in message:
        photo = message["photo"][-1]
        file_id, mime_type, file_name = photo["file_id"], "image/jpeg", "photo.jpg"

    if not file_id:
        await send_message(chat_id, "⚠️ 파일을 인식하지 못했습니다.")
        return

    try:
        await send_message(chat_id, f"📂 *{file_name}* 분석 중...")
        file_bytes = await download_file(file_id)
        file_block = await parse_file(file_bytes, mime_type, file_name)
        if is_task_request(caption):
            asyncio.create_task(run_team_pipeline(chat_id, caption, file_blocks=[file_block]))
        else:
            reply = await ask_daejang(caption, file_blocks=[file_block])
            await send_message(chat_id, reply)
    except Exception as e:
        logger.error(f"파일 처리 오류: {e}")
        await send_message(chat_id, f"⚠️ 파일 처리 오류: {str(e)}")


# ──────────────────────────────────────────
# 명령어 처리
# ──────────────────────────────────────────
async def handle_command(chat_id: int, text: str) -> bool:
    global stop_requested, interrupt_message, is_working
    global conversation_history, pending_task

    cmd = text.strip()

    # /stop
    if cmd == "/stop":
        if is_working:
            stop_requested = True
            await send_message(chat_id, "⛔ 중단 요청됨. 현재 단계 완료 후 중단됩니다...")
        else:
            await send_message(chat_id, "진행 중인 작업이 없습니다.")
        return True

    # 계속
    if cmd == "계속" and is_working:
        interrupt_message = "계속"
        return True

    # /start
    if cmd == "/start":
        await send_message(chat_id,
            "👋 *JW대장 v4 — 가동 중*\n\n"
            "━━ 📋 *팀 작업* ━━\n"
            "`○○ 포트폴리오 북 만들어줘`\n"
            "`2차완료 홍길동`\n"
            "`계약완료 홍길동`\n"
            "`카드뉴스 초안 - [주제]`\n"
            "`핀사이트랩스 리포트 - [주제]`\n\n"
            "━━ 📊 *섹터 분석* ━━\n"
            "`/분석 반도체` — 즉시 밸류에이션\n"
            "`/전체분석` — 설정된 전 섹터\n"
            "`/섹터목록` — 분석 가능 섹터\n\n"
            "━━ 🎛 *제어* ━━\n"
            "`/stop` — 작업 중단\n"
            "`계속` — 인터럽트 후 재개\n"
            "`확정` / `수정: [내용]` / `다시`\n"
            "`/clear` — 전체 초기화\n\n"
            "💬 자유 질문도 가능합니다"
        )
        return True

    # /clear
    if cmd == "/clear":
        conversation_history = []
        pending_task = None
        is_working = False
        stop_requested = False
        interrupt_message = None
        await send_message(chat_id, "🗑 전체 초기화 완료")
        return True

    # /help
    if cmd == "/help":
        await send_message(chat_id,
            "*JW대장 사용법*\n\n"
            "*팀 작업*\n"
            "`○○ 포트폴리오 북 만들어줘`\n"
            "`2차완료 [고객명]` / `계약완료 [고객명]`\n"
            "`카드뉴스 초안 - [주제]`\n\n"
            "*섹터 분석*\n"
            "`/분석 [섹터명]` — 예) `/분석 반도체`\n"
            "`/전체분석` — 전 섹터\n"
            "`/섹터목록` — 섹터명 확인\n\n"
            "*작업 제어*\n"
            "`/stop` — 중단 | `계속` — 재개\n"
            "`확정` / `수정: [내용]` / `다시`\n"
            "`/clear` — 초기화"
        )
        return True

    # /섹터목록
    if cmd == "/섹터목록":
        sectors = [
            "반도체", "IT하드웨어", "소프트웨어", "미디어·교육",
            "자동차", "자동차부품", "조선", "기계",
            "은행", "증권", "보험",
            "제약", "바이오", "헬스케어",
            "에너지", "화학", "철강", "건설",
            "음식료·담배", "유통", "화장품·의류"
        ]
        active = [f"• *{s}*" if s in ANALYSIS_SECTORS else f"• {s}" for s in sectors]
        await send_message(chat_id,
            "📋 *분석 가능 섹터 (WICS 기준)*\n"
            "굵게 표시 = 현재 설정 활성\n\n" +
            "\n".join(active) +
            "\n\n사용법: `/분석 반도체`"
        )
        return True

    # /분석 [섹터명]
    if cmd.startswith("/분석"):
        parts = cmd.split(maxsplit=1)
        if len(parts) < 2:
            await send_message(chat_id, "섹터명을 입력해주세요.\n예) `/분석 반도체`")
            return True
        sector = parts[1].strip()
        asyncio.create_task(run_sector_analysis(chat_id, sector))
        return True

    # /전체분석
    if cmd == "/전체분석":
        async def _all():
            for s in ANALYSIS_SECTORS:
                await run_sector_analysis(chat_id, s)
                await asyncio.sleep(3)
            await send_message(chat_id, "✅ 전체 섹터 분석 완료")
        asyncio.create_task(_all())
        return True

    return False


# ──────────────────────────────────────────
# Webhook
# ──────────────────────────────────────────
@app.post("/webhook")
async def webhook(request: Request):
    global interrupt_message, is_working

    data = await request.json()
    message = data.get("message")
    if not message:
        return {"ok": True}

    chat_id = message["chat"]["id"]
    if chat_id != ALLOWED_CHAT_ID:
        return {"ok": True}

    # 파일/이미지
    if "document" in message or "photo" in message:
        if is_working:
            file_name = message.get("document", {}).get("file_name", "파일")
            interrupt_message = f"[파일 첨부됨: {file_name}] {message.get('caption', '')}"
            await send_message(chat_id,
                f"⚡ 작업 중 파일 수신됨.\n현재 단계 완료 후 처리합니다.\n지금 중단: `/stop`")
        else:
            await handle_file_message(chat_id, message)
        return {"ok": True}

    text = message.get("text", "").strip()
    if not text:
        return {"ok": True}

    # 명령어 처리 (최우선)
    if await handle_command(chat_id, text):
        return {"ok": True}

    # 작업 중 → 인터럽트
    if is_working:
        interrupt_message = text
        await send_message(chat_id,
            f"⚡ *메시지 수신됨*\n`{text}`\n\n현재 단계 완료 후 처리합니다.\n지금 중단: `/stop`")
        return {"ok": True}

    # 컨펌 처리
    if await handle_confirm(chat_id, text):
        return {"ok": True}

    await send_typing(chat_id)

    try:
        if is_task_request(text):
            asyncio.create_task(run_team_pipeline(chat_id, text))
        else:
            reply = await ask_daejang(text)
            await send_message(chat_id, reply)
    except Exception as e:
        logger.error(f"오류: {e}")
        await send_message(chat_id, f"⚠️ 오류: {str(e)}")

    return {"ok": True}


@app.get("/")
async def health():
    return {"status": "JW대장 v4 운영중 ✅ — AGENTS.md 통합 완료"}
